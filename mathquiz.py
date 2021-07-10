from tkinter import *
from tkinter import messagebox
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
root = Tk()
root.geometry("600x400")
root.title("mathquiz")
var = StringVar()
question1=tk=StringVar()

wb=Workbook()
wb=load_workbook("Book1.xlsx")

def Loaddata():
    pass


def sumbit():
    button_choose_a.configure(bg="red")
    messagebox.askokcancel()
    root.destroy()
def Click_returnAnwser_A():
    button_choose_a.configure(bg="yellow")
def Click_returnAnwser_B():
    button_choose_b.configure(bg="yellow")
def Click_returnAnwser_C():
    button_choose_c.configure(bg="yellow")
def Click_returnAnwser_D():
    button_choose_d.configure(bg="yellow")


label_question = Label(root, text = ' question1: Nguyên thủ những nước nào sau đây tham dự Hội nghị Ianta (2/1945) ' , width=70 , borderwidth= 10 , font=('calibre',10,'normal'))
label_question.grid(row=0,column=0,columnspan=4,padx=10,pady=10)
button_choose_a= Button(root,text='A: Anh, Pháp, Mĩ.',padx=50,pady=20,width=15,command= Click_returnAnwser_A)
button_choose_b= Button(root,text='B: Anh, Mĩ, Liên Xô.',padx=50,pady=20,width=15,command= Click_returnAnwser_B)
button_choose_c= Button(root,text='C: Anh, Pháp, Đức.',padx=50,pady=20,width=15,command= Click_returnAnwser_C)
button_choose_d= Button(root,text='D: Mĩ, Liên Xô, Trung Quốc.',padx=50,pady=20,width=15,command= Click_returnAnwser_D)
button_next= Button(root,text='next',width=10,command = sumbit)
button_choose_a.grid(row=1,column=1)
button_choose_b.grid(row=1,column=2)
button_choose_c.grid(row=2,column=1)
button_choose_d.grid(row=2,column=2)
button_next.place(x=400,y=300)
root.mainloop()






